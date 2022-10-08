import requests
import pandas as pd
from openpyxl.styles import PatternFill

with open('url_list.txt', 'r') as f:
    url_list = f.read().splitlines()


headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.102 Safari/537.36"
}

status_description = {
    100: "Continue",
    101: "Switching Protocols",
    102: "Processing",
    103: "Early Hints",
    200: "OK",
    201: "Created",
    202: "Accepted",
    203: "Non-Authoritative Information",
    204: "No Content",
    205: "Reset Content",
    206: "Partial Content",
    207: "Multi-Status",
    208: "Already Reported",
    226: "IM Used",
    300: "Multiple Choices",
    301: "Moved Permanently",
    302: "Found",
    303: "See Other",
    304: "Not Modified",
    305: "Use Proxy",
    306: "Switch Proxy",
    307: "Temporary Redirect",
    308: "Permanent Redirect",
    400: "Bad Request",
    401: "Unauthorized",
    402: "Payment Required",
    403: "Forbidden",
    404: "Not Found",
    405: "Method Not Allowed",
    406: "Not Acceptable",
    407: "Proxy Authentication Required",
    408: "Request Timeout",
    409: "Conflict",
    410: "Gone",
    411: "Length Required",
    412: "Precondition Failed",
    413: "Payload Too Large",
    414: "URI Too Long",
    415: "Unsupported Media Type",
    416: "Range Not Satisfiable",
    417: "Expectation Failed",
    418: "I'm a teapot",
    421: "Misdirected Request",
    422: "Unprocessable Entity",
    423: "Locked",
    424: "Failed Dependency",
    425: "Too Early",
    426: "Upgrade Required",
    428: "Precondition Required",
    429: "Too Many Requests",
    431: "Request Header Fields Too Large",
    451: "Unavailable For Legal Reasons",
    500: "Internal Server Error",
    501: "Not Implemented",
    502: "Bad Gateway",
    503: "Service Unavailable",
    504: "Gateway Timeout",
    505: "HTTP Version Not Supported",
    506: "Variant Also Negotiates",
    507: "Insufficient Storage",
    508: "Loop Detected",
    510: "Not Extended",
    511: "Network Authentication Required",
}


result = []

try:
    for url in url_list:
        try:
            r = requests.get(url, headers=headers,allow_redirects=False)
            r.raise_for_status()
            if r.status_code in status_description:
                print("URL:", url)
                print("Status Code: {}".format(r.status_code))
                print("Status Description: {}".format(status_description[r.status_code]))
                result.append([url, r.status_code, status_description[r.status_code]])
        except Exception as e:
            print(e)
            result.append([url, e, e])

except requests.exceptions.HTTPError as err:
    print("Error: ", err)

df = pd.DataFrame(result, columns=["URL", "Status Code", "Status Description"])

writer = pd.ExcelWriter('output.xlsx', engine='openpyxl')

df.to_excel(writer, sheet_name='Sheet1', index=False)

wb = writer.book

ws = writer.sheets['Sheet1']


for i in range(1, len(df.index)+1):
    if str(df.iloc[i-1, 1])[0] == "1":
        ws.cell(row=i+1, column=4).value = "Informational"
    elif str(df.iloc[i-1, 1])[0] == "2":
        ws.cell(row=i+1, column=4).value = "Successful"
    elif str(df.iloc[i-1, 1])[0] == "3":
        ws.cell(row=i+1, column=4).value = "Redirection"
    elif str(df.iloc[i-1, 1])[0] == "4":
        ws.cell(row=i+1, column=4).value = "Client Error"
    elif str(df.iloc[i-1, 1])[0] == "5":
        ws.cell(row=i+1, column=4).value = "Server Error"


# add header for this new column
ws.cell(row=1, column=4).value = "Category"

# if category is "Client Error" or  "Server Error" or rediretion then fill the cell with red color. if category is Informational then fill the cell with yello color
for i in range(1, len(df.index)+1):
    if ws.cell(row=i+1, column=4).value == "Client Error" or ws.cell(row=i+1, column=4).value == "Server Error" or ws.cell(row=i+1, column=4).value == "Redirection":
        ws.cell(row=i+1, column=4).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    elif ws.cell(row=i+1, column=4).value == "Informational":
        ws.cell(row=i+1, column=4).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')


writer.close()